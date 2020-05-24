#! /usr/bin/env python
"""
Script to generate Network configuration files by 
combining data from excel files with Jinja templates 
"""


from openpyxl import load_workbook
from jinja2 import Template

templates_dir = "./templates"
outputcfg_dir = "./outputconfigs"

xwb_source_file = "ip_list_v0.xlsx"
access_template_file   = f"{templates_dir}/nx_int_access.j2"
uplink_template_file   = f"{templates_dir}/nx_int_uplink.j2"
mlag_template_file     = f"{templates_dir}/nx_vpc.j2"
vlan_template_file     = f"{templates_dir}/nx_vlans.j2"
fhrp_template_file     = f"{templates_dir}/nx_fhrp.j2"
portchep_template_file = f"{templates_dir}/nx_portch_endpoint.j2"
portchul_template_file = f"{templates_dir}/nx_portch_uplink.j2"
nxos_template_file     = f"{templates_dir}/nx_base.j2"
ospf_template_file     = f"{templates_dir}/nx_ospf.j2"
mgmt_template_file     = f"{templates_dir}/nx_mgmt.j2"

# Open up the Jinja template file (as text) and then create a Jinja Template Object 
with open(access_template_file) as f:
    access_template = Template(f.read(), keep_trailing_newline=True)

with open(uplink_template_file) as f:
    uplink_template = Template(f.read(), keep_trailing_newline=True)

with open(mlag_template_file) as f:
    mlag_template = Template(f.read(), keep_trailing_newline=True)

with open(vlan_template_file) as f:
    vlan_template = Template(f.read(), keep_trailing_newline=True)

with open(fhrp_template_file) as f:
    fhrp_template = Template(f.read(), keep_trailing_newline=True)

with open(portchul_template_file) as f:
    portchul_template = Template(f.read(), keep_trailing_newline=True)

with open(portchep_template_file) as f:
    portchep_template = Template(f.read(), keep_trailing_newline=True)

with open(nxos_template_file) as f:
    nxos_template = Template(f.read(), keep_trailing_newline=True)

with open(ospf_template_file) as f:
    ospf_template = Template(f.read(), keep_trailing_newline=True)

with open(mgmt_template_file) as f:
    mgmt_template = Template(f.read(), keep_trailing_newline=True)


# Open up the excel file containing the data 
xwb = load_workbook(xwb_source_file, data_only=True)
xsheet_sw = xwb["switchnames"]
xsheet_vl = xwb["vlans"]
xsheet_ip = xwb["ip_list"]
xsheet_ul = xwb["sw_uplinks"]
xsheet_ospf = xwb["ospf"]

#extract columns from switchnames sheet
hstnm_sw = xsheet_sw['B']
obmad_sw = xsheet_sw['C']
obmmk_sw = xsheet_sw['D']
obmgw_sw = xsheet_sw['E']
obmvf_sw = xsheet_sw['F']
ibmad_sw = xsheet_sw['G']
ibmmk_sw = xsheet_sw['H']
ibmgw_sw = xsheet_sw['I']
ibmvf_sw = xsheet_sw['J']
domnm_sw = xsheet_sw['K']
mlgdm_sw = xsheet_sw['L']
mlgpr_sw = xsheet_sw['M']
mlgsr_sw = xsheet_sw['N']
mlgdt_sw = xsheet_sw['O']
mlgvf_sw = xsheet_sw['P']


#extract columns from vlan sheet
vlnum_vl = xsheet_vl['B']
vlnam_vl = xsheet_vl['C']
vlnet_vl = xsheet_vl['D']
vlmsk_vl = xsheet_vl['E']
vlbit_vl = xsheet_vl['F']
vlgtw_vl = xsheet_vl['G']
fhsw1_vl = xsheet_vl['H']
fhpr1_vl = xsheet_vl['I']
fhad1_vl = xsheet_vl['J']
fhsw2_vl = xsheet_vl['K']
fhpr2_vl = xsheet_vl['L']
fhad2_vl = xsheet_vl['M']
fhprt_vl = xsheet_vl['N']
fhvrf_vl = xsheet_vl['O']
fhigp_vl = xsheet_vl['P']
fhmtu_vl = xsheet_vl['Q']

# extract columns from IP_List sheet for endpoint interfaces
hstnm_ep = xsheet_ip['D']
hstpt_ep = xsheet_ip['E']
prpse_ep = xsheet_ip['F']
swtch_ep = xsheet_ip['G']
intid_ep = xsheet_ip['H']
intno_ep = xsheet_ip['I']
prtch_ep = xsheet_ip['J']
mlgid_ep = xsheet_ip['K']
intrl_ep = xsheet_ip['L']
vlnid_ep = xsheet_ip['M']
swpad_ep = xsheet_ip['N']
swpmk_ep = xsheet_ip['O']
swigp_ep = xsheet_ip['P']
swmtu_ep = xsheet_ip['Q']

# extract columns from Uplink sheet for switch to switch interfaces
srcsw_ul = xsheet_ul['B']
srpid_ul = xsheet_ul['C']
srpno_ul = xsheet_ul['D']
prtch_ul = xsheet_ul['E']
mlgid_ul = xsheet_ul['F']
intrl_ul = xsheet_ul['G']
vlnid_ul = xsheet_ul['H']
swpad_ul = xsheet_ul['I']
swpmk_ul = xsheet_ul['J']
spvrf_ul = xsheet_ul['K']
swigp_ul = xsheet_ul['L']
swmtu_ul = xsheet_ul['M']
dstsw_ul = xsheet_ul['N']
dspid_ul = xsheet_ul['O']
dspno_ul = xsheet_ul['P']


# extract columns from ospf sheet for routing
swtch_ospf = xsheet_ospf['B']
loopb_opsf = xsheet_ospf['C']
rtrid_ospf = xsheet_ospf['D']
araid_ospf = xsheet_ospf['E']
prcid_ospf = xsheet_ospf['F']
aukey_ospf = xsheet_ospf['G']

#define j2 template as a function for access interfaces
def access_generate (int_ty, int_no, vln_no, hst_nm, hst_pt, int_pr, prt_ch, int_rl,
swp_ad, swp_mk, igp_pt, mtu_pt):
    access_config = access_template.render(
        inttype  = int_ty,
        int_num  = int_no,
        vlan     = vln_no,
        hostname = hst_nm,
        link     = hst_pt,
        purpose  = int_pr,
        portch   = prt_ch,
        introle  = int_rl,
        ipaddr   = swp_ad,
        subnet   = swp_mk,
        )
    return(access_config)

#define j2 template as a function for uplink interfaces
def uplink_generate (srp_ty, srp_no, prt_ch, int_rl, vln_no, swp_ad, swp_mk, vrf_pt,
igp_pt, mtu_pt, dsw_nm, dsp_ty, dsp_no):
    uplink_config = uplink_template.render(
        inttype    = srp_ty,
        int_num    = srp_no,
        portch     = prt_ch,
        introle    = int_rl,
        vlanid     = vln_no,
        ipaddr     = swp_ad,
        subnet     = swp_mk,
        intfvrf    = vrf_pt,
        igpprot    = igp_pt,
        portmtu    = mtu_pt,
        destsw     = dsw_nm,
        dinttype   = dsp_ty,
        dintno     = dsp_no,
        ospfprocid = prcid_ospf[1].value,
        ospfareaid = araid_ospf[1].value,
        ospfautkey = aukey_ospf[1].value,
        )
    return(uplink_config)

#define j2 template as a function for mlag configs
def mlag_generate (vpc_dm, vpc_pr, kpa_sr, kpa_dt, kpa_vf):
    vpc_config = mlag_template.render(
        vpcdomain      = vpc_dm,
        vpcpriority    = vpc_pr,
        keepalivesrce  = kpa_sr,
        keepalivedest  = kpa_dt,
        keepalivevrf   = kpa_vf,
    )
    return(vpc_config)

#define j2 template as a function for endpoint portchannel interfaces
def portchep_generate (pch_id, mlg_id, vln_no, int_rl, hst_nm):
    portchep_config = portchep_template.render(
        portchid    = pch_id,
        mlagid      = mlg_id,
        vlanid      = vln_no,
        introle     = int_rl,
        hostname    = hst_nm,
    )
    return(portchep_config)

def portchul_generate (pch_id, mlg_id, vln_no, int_rl, dst_sw):
    portchul_config = portchul_template.render(
        portchid    = pch_id,
        mlagid      = mlg_id,
        vlanid      = vln_no,
        introle     = int_rl,
        destsw      = dst_sw,
    )
    return(portchul_config)

#define j2 template as a function for vlans
def vlan_generate (vln_no, vln_nm):
    vlan_gen= vlan_template.render(
        vlanid   = vln_no,
        vlanname = vln_nm,
        )
    return(vlan_gen)

#define j2 template as a function for fhrp
def fhrp_generate (vln_no, vln_nm, fhr_ad, fhr_nt, vln_gw, fhr_pr, fhr_vf, fhr_pt,
igp_pt, mtu_pt):
    fhrp_gen= fhrp_template.render(
        vlanid          = vln_no,
        vlandescription = vln_nm,
        fhrpipaddress   = fhr_ad,
        bitmask         = fhr_nt,
        fhrpgroup       = vln_no,
        vlangateway     = vln_gw,
        fhrppriority    = fhr_pr,
        interfacevrf    = fhr_vf,
        fhrpprotocol    = fhr_pt,
        fhrpigp         = igp_pt,
        fhrpmtu         = mtu_pt,
        ospfprocid      = prcid_ospf[1].value,
        ospfareaid      = araid_ospf[1].value,
        )
    return(fhrp_gen)

#define j2 template as a function for ospf
def ospf_generate (prc_id, rtr_id):
    ospf_config = ospf_template.render(
        processid = prc_id,
        routerid  = rtr_id,
    )
    return(ospf_config)

#define j2 template as a function for switch management configs
def mgmt_generate (obm_vf, obm_ad, obm_mk, obm_gw):
    mgmt_config = mgmt_template.render(
        obmgmtvrf = obm_vf,
        obipaddr  = obm_ad,
        obsubnet  = obm_mk,
        obgateway = obm_gw,
    )
    return(mgmt_config)

#define j2 template as a function to save switches final configs
def save_config (swc_nm, dmn_nm, acc_cf, pac_cf, upl_cf, pul_cf, mlg_cf, vln_cf, fhr_cf, osp_cf, mgt_cf):
    save_config = nxos_template.render(
        hostname       = swc_nm,
        domainname     = dmn_nm,
        intf_access    = acc_cf,
        intf_uplink    = upl_cf,
        vpcconfigs     = mlg_cf,
        vlanconfigs    = vln_cf,
        fhrpconfigs    = fhr_cf,
        portchaccess   = pac_cf,
        portchuplink   = pul_cf,
        ospfconfigs    = osp_cf,
        mgmtconfigs    = mgt_cf,
    )
    #open file and save configuration with switch hostname
    with open(f"{outputcfg_dir}/{swc_nm}" + "config.txt", "w") as f:
        f.write(save_config)


#Main function to generate the configurations for all the templates
for x in range (1, xsheet_sw.max_row):
    access_configs   = ""
    uplink_configs   = ""
    mlag_configs     = ""
    portchep_configs = ""
    portchul_configs = ""
    portchep_check   = set() #duplicate check for portchannel
    portchul_check   = set() #duplicate check for portchannel
    vlanid_configs   = ""
    fhrp_configs     = ""
    ospf_configs     = ""
    mgmt_configs     = ""

    #generate management configuration of the switch
    mgmt_configs = mgmt_generate (obmvf_sw[x].value, obmad_sw[x].value, obmmk_sw[x].value, obmgw_sw[x].value)
    
    #generate vpc/mlag configurations
    if mlgdm_sw[x].value != None:
        mlag_configs = mlag_generate(mlgdm_sw[x].value, mlgpr_sw[x].value, mlgsr_sw[x].value, mlgdt_sw[x].value,
        mlgvf_sw[x].value)

    #Loop through IP list sheet 
    for y in range (1, xsheet_ip.max_row):
        if swtch_ep[y].value == hstnm_sw[x].value:
            # find the switchname in the call and call access config function for it to connected endpoint
            access_config = access_generate(intid_ep[y].value, intno_ep[y].value, vlnid_ep[y].value,
            hstnm_ep[y].value, hstpt_ep[y].value, prpse_ep[y].value, prtch_ep[y].value, intrl_ep[y].value,
            swpad_ep[y].value, swpmk_ep[y].value, swigp_ep[y].value, swmtu_ep[y].value)

            # Append interface configurations
            access_configs += access_config

            #check if endpoint portchannel is configured, if yes then generate porchannel configurations
            if prtch_ep[y].value != None and prtch_ep[y].value not in portchep_check:
                portchep_config = portchep_generate(prtch_ep[y].value, mlgid_ep[y].value,vlnid_ep[y].value,
                intrl_ep[y].value, hstnm_ep[y].value)
                portchep_configs += portchep_config
                
                #update set value to avoid duplicate portchannel interfaces
                portchep_check.add(prtch_ep[y].value)

    #Loop through Switch uplink sheet
    for y in range (1, xsheet_ul.max_row):
        if srcsw_ul[y].value == hstnm_sw[x].value:
            # find the switchname in the call and call uplink config function for it to connected endpoint
            uplink_config = uplink_generate(srpid_ul[y].value, srpno_ul[y].value, prtch_ul[y].value,
            intrl_ul[y].value, vlnid_ul[y].value, swpad_ul[y].value, swpmk_ul[y].value, spvrf_ul[y].value,
            swigp_ul[y].value, swmtu_ul[y].value, dstsw_ul[y].value, dspid_ul[y].value,dspno_ul[y].value)

            # Append interface configuration
            uplink_configs += uplink_config

            #check if uplink portchannel is configured, if yes then generate porchannel configurations
            if prtch_ul[y].value != None and prtch_ul[y].value not in portchul_check:
                portchul_config = portchul_generate(prtch_ul[y].value, mlgid_ul[y].value, vlnid_ul[y].value,
                intrl_ul[y].value, dstsw_ul[y].value)
                portchul_configs += portchul_config

                #update set value to avoid duplicate portchannel interfaces
                portchul_check.add(prtch_ul[y].value)

    #generate vlan and fhrp configurations        
    for y in range (1, xsheet_vl.max_row):
        #generate vlan configuration
        if swtch_ep[y].value == hstnm_sw[x].value:
            vlan_config = vlan_generate (vlnum_vl[y].value, vlnam_vl[y].value)

            #append vlan configurations together
            vlanid_configs += vlan_config

        #generate fhrp configuration
        if fhsw1_vl[y].value == hstnm_sw[x].value:
            fhrp_config = fhrp_generate(vlnum_vl[y].value, vlnam_vl[y].value, fhad1_vl[y].value, vlbit_vl[y].value,
            vlgtw_vl[y].value, fhpr1_vl[y].value, fhvrf_vl[y].value, fhprt_vl[y].value, fhigp_vl[y].value, fhmtu_vl[y].value)

            #append fhrp1 configuration
            fhrp_configs += fhrp_config

        elif fhsw2_vl[y].value == hstnm_sw[x].value:
            fhrp_config = fhrp_generate(vlnum_vl[y].value, vlnam_vl[y].value, fhad2_vl[y].value, vlbit_vl[y].value,
            vlgtw_vl[y].value, fhpr2_vl[y].value, fhvrf_vl[y].value, fhprt_vl[y].value, fhigp_vl[y].value, fhmtu_vl[y].value)

            #append fhrp2 configuration
            fhrp_configs += fhrp_config

    #generate ospf configurations
    for y in range (1, xsheet_ospf.max_row):
        if swtch_ospf[y].value == hstnm_sw[x].value:
            ospf_config = ospf_generate (prcid_ospf[y].value, rtrid_ospf[y].value)
            
            # Append ospf configurations together
            ospf_configs += ospf_config

    #call save_config to save configuration for a switch
    save_config(hstnm_sw[x].value, domnm_sw[x].value, access_configs, portchep_configs, uplink_configs, portchul_configs, mlag_configs,
    vlanid_configs, fhrp_configs, ospf_configs, mgmt_configs)